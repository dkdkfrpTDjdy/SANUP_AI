# %% [cell 1 type=markdown]
# # 임대대수 대시보드 전체 YYMM RAW 자동 반영 버전
# 
# - 입력 파일에서 `YYMM RAW` 시트를 자동 탐지합니다.
# - 직전 월과의 체인으로 `연장 여부`를 계산합니다.
# - 대시보드/팀별/개인별/DEBUG 시트에 탐지된 모든 월을 반영합니다.

# %% [cell 2 type=code]
# -*- coding: utf-8 -*-
"""
YYMM RAW 전체 월 자동 탐지 버전
- 입력 파일에서 '#### RAW' 형식의 시트를 모두 찾아 월순으로 정렬
- 첫 월은 자체 기준으로 신규/기존만 산출
- 둘째 월부터는 직전 월과 코드 기준으로 연장 여부를 체인 계산
- Dash Board 피벗 / 팀별 / 개인별 / DEBUG 시트에 모든 YYMM RAW 월을 반영
"""
import re
import math
import json
import warnings
from typing import Dict, List, Tuple

import numpy as np
import pandas as pd
from pandas.tseries.offsets import MonthEnd
from openpyxl import Workbook

warnings.filterwarnings("ignore", category=UserWarning)

# %% [cell 3 type=markdown]
# ## 0. 실행 변수

# %% [cell 4 type=code]
IN_PATH = r"D:\works\데이터분석\산업장비\AI\26년_3월_임대대수_데이터.xlsx"
OUT_PATH = r"D:\works\데이터분석\산업장비\AI\output_debug_dashboard_all_raw.xlsx"
REF_SHEET = "기준정보"

DEFAULT_INCLUDED_ASSET_STATUS = {"배송완료", "출고작업완료"}

BLOCK1_ITEMS = ["AWP", "FL", "FL물류", "핸드파렛트 트럭", "청소기기", "발전기", "COMP", "FL 기타"]
BLOCK2_ITEMS = ["AWP", "FL", "FL물류", "핸드파렛트 트럭", "청소기기", "발전기", "COMP"]

EXPECTED_COLUMNS = [
    "최초시작일", "청구종료일", "자산상태", "자산구분",
    "영업팀", "영업담당자", "대당 월렌탈료", "취득가",
    "자산번호", "BP번호", "주문번호", "1"
]

RAW_SHEET_PATTERN = re.compile(r"^(\d{4})\s*RAW$")

# %% [cell 5 type=markdown]
# ## 1. 공통 유틸

# %% [cell 6 type=code]
def print_title(title: str):
    print("\n" + "=" * 100)
    print(title)
    print("=" * 100)

def normalize_text(x) -> str:
    if pd.isna(x):
        return ""
    s = str(x)
    s = s.replace("\u00A0", " ")
    s = s.replace("\n", " ")
    s = s.replace("\r", " ")
    s = re.sub(r"\s+", " ", s)
    return s.strip()

def normalize_colname(c) -> str:
    if pd.isna(c):
        return ""
    s = normalize_text(c)
    if s in ("1.0", "1"):
        return "1"
    return s

def normalize_colnames(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out.columns = [normalize_colname(c) for c in out.columns]
    return out

def coalesce_columns(df: pd.DataFrame) -> pd.DataFrame:
    cols = list(df.columns)
    dupes = [c for c in pd.Index(cols).unique().tolist() if cols.count(c) > 1]
    if not dupes:
        return df

    out = pd.DataFrame(index=df.index)
    handled = set()
    for c in cols:
        if c in handled:
            continue
        same_cols = [x for x in cols if x == c]
        if len(same_cols) == 1:
            out[c] = df[c]
        else:
            temp = df.loc[:, [col for col in df.columns if col == c]].copy()
            merged = temp.iloc[:, 0]
            for i in range(1, temp.shape[1]):
                merged = merged.combine_first(temp.iloc[:, i])
            out[c] = merged
        handled.add(c)
    return out

def safe_divide(a, b):
    if b is None or pd.isna(b) or b == 0:
        return np.nan
    if a is None or pd.isna(a):
        return np.nan
    return a / b

def month_end_from_yymm(yymm: str) -> pd.Timestamp:
    yy = int(yymm[:2])
    mm = int(yymm[2:])
    year = 2000 + yy
    return pd.Timestamp(year=year, month=mm, day=1) + MonthEnd(1)

def month_label_from_yymm(yymm: str) -> str:
    return f"{int(yymm[2:])}월"

def build_month_label_map(months: List[str]) -> Dict[str, str]:
    simple = [month_label_from_yymm(m) for m in months]
    if len(set(simple)) == len(simple):
        return {m: month_label_from_yymm(m) for m in months}
    return {m: f"{2000 + int(m[:2])}.{int(m[2:]):02d}" for m in months}

def num_or_none(v):
    if pd.isna(v):
        return None
    if isinstance(v, (np.integer, int)):
        return int(v)
    if isinstance(v, (np.floating, float)):
        if math.isnan(v):
            return None
        return float(v)
    return v

def collapse_code_status(df: pd.DataFrame, code_col="코드", status_col="연장대상구분") -> pd.Series:
    temp = df[[code_col, status_col]].copy()
    temp[code_col] = temp[code_col].map(normalize_text)
    temp[status_col] = temp[status_col].map(normalize_text)

    def reducer(s: pd.Series) -> str:
        vals = set(s.dropna().astype(str))
        return "연장대상" if "연장대상" in vals else "정상"

    return temp.groupby(code_col)[status_col].apply(reducer)

def find_all_raw_yymm_sheets(in_path: str) -> List[str]:
    xls = pd.ExcelFile(in_path)
    months = []
    for sheet_name in xls.sheet_names:
        m = RAW_SHEET_PATTERN.match(normalize_text(sheet_name))
        if m:
            months.append(m.group(1))

    months = sorted(set(months), key=lambda x: month_end_from_yymm(x))
    if not months:
        raise ValueError("입력 파일에서 'YYMM RAW' 형식의 시트를 찾지 못했습니다.")
    return months

# %% [cell 7 type=markdown]
# ## 2. 날짜 / 숫자 파싱

# %% [cell 8 type=code]
def excel_date_to_timestamp_series(s: pd.Series) -> pd.Series:
    if pd.api.types.is_datetime64_any_dtype(s):
        return pd.to_datetime(s, errors="coerce")

    out = pd.Series(pd.NaT, index=s.index, dtype="datetime64[ns]")
    num = pd.to_numeric(s, errors="coerce")
    num_mask = num.notna()

    if num_mask.any():
        out.loc[num_mask] = pd.to_datetime(num.loc[num_mask], unit="D", origin="1899-12-30", errors="coerce")

    str_mask = ~num_mask
    if str_mask.any():
        out.loc[str_mask] = pd.to_datetime(s.loc[str_mask], errors="coerce")

    return out

def parse_numeric_series(s: pd.Series) -> pd.Series:
    s2 = s.astype(str)
    s2 = s2.str.replace(",", "", regex=False)
    s2 = s2.str.replace(" ", "", regex=False)
    s2 = s2.str.replace("\u00A0", "", regex=False)
    s2 = s2.replace({"": np.nan, "nan": np.nan, "None": np.nan, "-": np.nan})
    return pd.to_numeric(s2, errors="coerce")

# %% [cell 9 type=markdown]
# ## 3. 기준정보 로드

# %% [cell 10 type=code]
def load_reference_sheet(in_path: str, ref_sheet: str) -> pd.DataFrame:
    ref = pd.read_excel(in_path, sheet_name=ref_sheet, header=None)
    ref = ref.fillna("")
    ref = ref.map(normalize_text, na_action="ignore")
    return ref

def find_cell_positions(ref: pd.DataFrame, keyword: str) -> List[Tuple[int, int]]:
    hits = []
    for r in range(ref.shape[0]):
        for c in range(ref.shape[1]):
            if normalize_text(ref.iat[r, c]) == normalize_text(keyword):
                hits.append((r, c))
    return hits

def extract_vertical_pairs_below(ref: pd.DataFrame, start_row: int, key_col: int, val_col: int, stop_blank_streak: int = 3):
    pairs = []
    blank_streak = 0
    for r in range(start_row, ref.shape[0]):
        k = normalize_text(ref.iat[r, key_col]) if key_col < ref.shape[1] else ""
        v = normalize_text(ref.iat[r, val_col]) if val_col < ref.shape[1] else ""

        if k == "" and v == "":
            blank_streak += 1
            if blank_streak >= stop_blank_streak:
                break
            continue

        blank_streak = 0
        if k != "":
            pairs.append((k, v))
    return pairs

def extract_list_below(ref: pd.DataFrame, start_row: int, key_col: int, stop_blank_streak: int = 3):
    vals = []
    blank_streak = 0
    for r in range(start_row, ref.shape[0]):
        v = normalize_text(ref.iat[r, key_col]) if key_col < ref.shape[1] else ""
        if v == "":
            blank_streak += 1
            if blank_streak >= stop_blank_streak:
                break
            continue
        blank_streak = 0
        vals.append(v)
    return vals

def load_reference_maps(in_path: str, ref_sheet: str) -> Dict[str, object]:
    ref = load_reference_sheet(in_path, ref_sheet)

    item_map = {}
    hits = find_cell_positions(ref, "아이템 구분")
    if hits:
        r, c = hits[0]
        pairs = extract_vertical_pairs_below(ref, start_row=r + 1, key_col=c, val_col=c + 1)
        item_map = {normalize_text(k): normalize_text(v) for k, v in pairs if normalize_text(k)}

    fl_logistics_people = set()
    hits = find_cell_positions(ref, "FL 물류 구분")
    if hits:
        r, c = hits[0]
        vals = extract_list_below(ref, start_row=r + 1, key_col=c)
        fl_logistics_people = set(normalize_text(x) for x in vals if normalize_text(x))

    included_asset_status = set(DEFAULT_INCLUDED_ASSET_STATUS)
    hits = find_cell_positions(ref, "자산상태 구분")
    if hits:
        r, c = hits[0]
        pairs = extract_vertical_pairs_below(ref, start_row=r + 1, key_col=c, val_col=c + 1)
        temp = []
        for k, v in pairs:
            kk = normalize_text(k)
            vv = normalize_text(v)
            if kk and (vv in ("1", "Y", "y", "포함", "TRUE", "True", "사용", "대상") or vv == ""):
                temp.append(kk)
        if temp:
            included_asset_status = set(temp)

    print_title("기준정보 로딩 결과")
    print(f"아이템 구분 개수: {len(item_map)}")
    print(f"FL 물류 구분 개수: {len(fl_logistics_people)}")
    print(f"포함 자산상태: {sorted(included_asset_status)}")

    return {
        "item_map": item_map,
        "fl_logistics_people": fl_logistics_people,
        "included_asset_status": included_asset_status,
    }

# %% [cell 11 type=markdown]
# ## 4. RAW 로드 / 검증

# %% [cell 12 type=code]
def load_raw_sheet(in_path: str, sheet_name: str) -> pd.DataFrame:
    df = pd.read_excel(in_path, sheet_name=sheet_name)
    df = normalize_colnames(df)
    df = coalesce_columns(df)
    return df

def validate_raw_sheet(df: pd.DataFrame, sheet_name: str):
    print_title(f"{sheet_name} 시트 로딩 점검")
    print(f"{sheet_name}.shape = {df.shape}")
    print(df.head(3).to_string())

    for c in EXPECTED_COLUMNS:
        print(f" - {c}: {'OK' if c in df.columns else 'MISSING'}")

    missing = [c for c in EXPECTED_COLUMNS if c not in df.columns]
    if missing:
        raise ValueError(f"[{sheet_name}] 필수 컬럼 누락: {missing}")

# %% [cell 13 type=markdown]
# ## 5. 연장키 / 자산구분 표준화

# %% [cell 14 type=code]
def make_code(df: pd.DataFrame) -> pd.Series:
    a = df["자산번호"].map(normalize_text)
    b = df["BP번호"].map(normalize_text)
    c = df["주문번호"].map(normalize_text)
    return a + "-" + b + "-" + c

def make_extension_key(df: pd.DataFrame) -> pd.Series:
    a = df["자산번호"].map(normalize_text)
    b = df["BP번호"].map(normalize_text)
    return a + "-" + b

def normalize_asset_type_value(x: str) -> str:
    s = normalize_text(x)
    if s == "":
        return "미분류"

    mapping = {
        "당사자산": "당사자산",
        "당사 자산": "당사자산",
        "당사": "당사자산",
        "자사자산": "당사자산",
        "임차자산": "임차자산",
        "임차 자산": "임차자산",
        "임차": "임차자산",
        "렌탈용악세사리": "렌탈용악세사리",
    }
    return mapping.get(s, "미분류")

def normalize_asset_type_series(s: pd.Series) -> pd.Series:
    return s.map(normalize_asset_type_value)

def build_asset_type_debug_tables(df: pd.DataFrame, yymm: str):
    raw_dist = (
        df["자산구분"].fillna("").map(normalize_text).replace("", "<<빈값>>")
        .value_counts(dropna=False).rename_axis("자산구분_원천").reset_index(name="건수")
    )
    raw_dist.insert(0, "대상월", yymm)

    norm_dist = (
        df["자산구분_std"].fillna("미분류")
        .value_counts(dropna=False).rename_axis("자산구분_표준화").reset_index(name="건수")
    )
    norm_dist.insert(0, "대상월", yymm)

    sample = df[["자산번호", "BP번호", "주문번호", "코드", "연장키", "자산구분", "자산구분_std"]].copy()
    sample.insert(0, "대상월", yymm)
    sample = sample[sample["자산구분_std"] == "미분류"].head(200)

    return raw_dist, norm_dist, sample

# %% [cell 15 type=markdown]
# ## 6. 전처리

# %% [cell 16 type=code]
def diagnose_target_flag(df: pd.DataFrame, yymm: str, included_asset_status: set):
    print_title(f"{yymm} 대상여부 점검")
    print(df["대상여부"].value_counts(dropna=False).to_string())

def diagnose_item_mapping(df: pd.DataFrame, yymm: str):
    print_title(f"{yymm} 아이템 매핑 점검")
    print(f"아이템 NaN 비율: {df['아이템'].isna().mean():.4%}")

def diagnose_numeric_parsing(df: pd.DataFrame, yymm: str):
    print_title(f"{yymm} 숫자 파싱 점검")
    print(f"렌탈료 NaN 비율: {df['렌탈료'].isna().mean():.4%}")
    print(f"취득가 NaN 비율: {df['취득가_num'].isna().mean():.4%}")

def diagnose_asset_type(df: pd.DataFrame, yymm: str):
    print_title(f"{yymm} 자산구분 표준화 점검")
    print(df["자산구분_std"].value_counts(dropna=False).to_string())

def preprocess_month_raw(raw_df: pd.DataFrame, yymm: str, ref_maps: Dict[str, object]):
    df = raw_df.copy()

    당월기준 = month_end_from_yymm(yymm)
    신규기준월 = month_end_from_yymm(yymm) - MonthEnd(1)
    연장기준 = month_end_from_yymm(yymm) + MonthEnd(1)

    df["최초시작일"] = excel_date_to_timestamp_series(df["최초시작일"])
    df["청구종료일"] = excel_date_to_timestamp_series(df["청구종료일"])

    df["당월기준"] = 당월기준
    df["신규 기준월"] = 신규기준월
    df["연장기준"] = 연장기준

    df["자산구분_std"] = normalize_asset_type_series(df["자산구분"])

    included_asset_status = ref_maps["included_asset_status"]
    cond = (
        (df["최초시작일"] <= df["당월기준"]) &
        (df["자산구분_std"] != "렌탈용악세사리") &
        (df["자산상태"].map(normalize_text).isin(included_asset_status))
    )
    df["대상여부"] = np.where(cond, 1, 0)

    df["담당"] = (
        df["영업팀"].fillna("").astype(str).map(normalize_text) + " " +
        df["영업담당자"].fillna("").astype(str).map(normalize_text)
    ).str.strip()

    item_map = {normalize_text(k): normalize_text(v) for k, v in ref_maps["item_map"].items()}
    item_key = df["1"].fillna("").astype(str).map(normalize_text)
    df["아이템"] = item_key.map(item_map)
    df["미매핑_아이템키"] = np.where(df["아이템"].isna(), item_key, "")

    fl_logistics_people = set(normalize_text(x) for x in ref_maps["fl_logistics_people"])
    df["아이템 세분화"] = np.where(
        (df["아이템"] == "FL") & (df["담당"].map(normalize_text).isin(fl_logistics_people)),
        "FL물류",
        df["아이템"]
    )
    df["아이템 세분화"] = df["아이템 세분화"].fillna("미분류")

    df["렌탈료"] = parse_numeric_series(df["대당 월렌탈료"])
    df["취득가_num"] = parse_numeric_series(df["취득가"])

    df["연장대상구분"] = np.where(df["청구종료일"] <= df["연장기준"], "연장대상", "정상")
    df["코드"] = make_code(df)
    df["연장키"] = make_extension_key(df)

    df["재계약 여부"] = "-"
    df["연장 여부"] = "-"
    df["신규여부"] = "-"

    diagnose_target_flag(df, yymm, included_asset_status)
    diagnose_item_mapping(df, yymm)
    diagnose_numeric_parsing(df, yymm)
    diagnose_asset_type(df, yymm)

    asset_raw_dist, asset_norm_dist, asset_unknown_sample = build_asset_type_debug_tables(df, yymm)

    stats = {
        "rows": len(df),
        "cols": len(df.columns),
        "대상여부_1_count": int((df["대상여부"] == 1).sum()),
        "아이템_nan_ratio": float(df["아이템"].isna().mean()),
        "렌탈료_nan_ratio": float(df["렌탈료"].isna().mean()),
        "취득가_nan_ratio": float(df["취득가_num"].isna().mean()),
        "코드_중복건수": int(df["코드"].duplicated().sum()),
        "연장키_중복건수": int(df["연장키"].duplicated().sum()),
        "아이템세분화_top20": df["아이템 세분화"].value_counts().head(20).to_dict(),
        "자산구분_std_분포": df["자산구분_std"].value_counts(dropna=False).to_dict(),
    }

    debug_tables = {
        "asset_raw_dist": asset_raw_dist,
        "asset_norm_dist": asset_norm_dist,
        "asset_unknown_sample": asset_unknown_sample,
    }
    return df, stats, debug_tables

def initialize_first_month(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    out["재계약 여부"] = "-"
    out["연장 여부"] = "제외"
    out["신규여부"] = np.where(
        (out["신규 기준월"] - out["최초시작일"]).dt.days < 0,
        "신규",
        "기존"
    )
    out.loc[out["최초시작일"].isna(), "신규여부"] = "-"
    return out

# %% [cell 17 type=markdown]
# ## 7. 연장 체인

# %% [cell 18 type=code]
def build_extension_debug(prev_df: pd.DataFrame, out_df: pd.DataFrame, target_yymm: str) -> pd.DataFrame:
    prev_keys = prev_df[["코드", "연장대상구분"]].copy().rename(
        columns={"연장대상구분": "전월_연장대상구분"}
    )
    prev_keys = prev_keys.drop_duplicates(subset=["코드"])

    debug = out_df[[
        "코드", "연장키", "자산번호", "BP번호", "주문번호",
        "청구종료일", "연장대상구분", "재계약 여부", "연장 여부", "신규여부"
    ]].copy()
    debug.insert(0, "대상월", target_yymm)

    debug = debug.merge(prev_keys, on="코드", how="left")

    debug["매칭여부"] = np.where(debug["전월_연장대상구분"].notna(), "매칭", "매칭실패")
    debug["연장 여부 원인"] = np.select(
        [
            debug["매칭여부"] == "매칭실패",
            debug["재계약 여부"] == "연장대상",
            debug["재계약 여부"] == "정상"
        ],
        [
            "전월 코드 미매칭",
            "전월 연장대상",
            "전월 정상"
        ],
        default="기타"
    )
    return debug

def apply_extension_chain(prev_df: pd.DataFrame, target_df: pd.DataFrame, target_yymm: str):
    prev_reduce = collapse_code_status(prev_df, code_col="코드", status_col="연장대상구분")

    out = target_df.copy()
    out["재계약 여부"] = out["코드"].map(prev_reduce).fillna("-")
    out["연장 여부"] = np.where(out["재계약 여부"] == "연장대상", "연장", "제외")
    out["신규여부"] = np.where(
        out["연장 여부"] == "연장",
        "연장",
        np.where((out["신규 기준월"] - out["최초시작일"]).dt.days < 0, "신규", "기존")
    )
    out.loc[out["최초시작일"].isna(), "신규여부"] = "-"

    match_rate = float((out["재계약 여부"] != "-").mean())

    ext_check = out[["코드", "연장키", "연장대상구분", "재계약 여부", "연장 여부"]].copy()
    ext_check.insert(0, "대상월", target_yymm)
    ext_check.rename(columns={
        "연장대상구분": "당월_연장대상구분",
        "재계약 여부": "전월_연장대상구분_lookup",
        "연장 여부": "당월_연장 여부",
    }, inplace=True)

    ext_debug = build_extension_debug(prev_df, out, target_yymm)
    return out, ext_check, ext_debug, match_rate

# %% [cell 19 type=markdown]
# ## 8. 대시보드 집계

# %% [cell 20 type=code]
def summarize_item_month(df: pd.DataFrame, item: str) -> Dict[str, float]:
    base = df[df["대상여부"] == 1].copy()
    item_df = base[base["아이템 세분화"] == item].copy()

    existing_df = item_df[item_df["신규여부"] == "기존"]
    new_df = item_df[item_df["신규여부"] == "신규"]
    ext_df = item_df[item_df["신규여부"] == "연장"]
    ext_target_df = item_df[item_df["연장대상구분"] == "연장대상"]

    own_df = item_df[item_df["자산구분_std"] == "당사자산"]
    leased_df = item_df[item_df["자산구분_std"] == "임차자산"]
    unknown_df = item_df[item_df["자산구분_std"] == "미분류"]

    total_rental = item_df["렌탈료"].sum(min_count=1) if not item_df.empty else 0
    total_cost = item_df["취득가_num"].sum(min_count=1) if not item_df.empty else 0
    recovery_rate = safe_divide(total_rental, total_cost)

    own_cnt = len(own_df)
    leased_cnt = len(leased_df)
    asset_cnt = own_cnt + leased_cnt

    return {
        "렌탈료": total_rental,
        "렌탈료_기존": existing_df["렌탈료"].sum(min_count=1) if not existing_df.empty else 0,
        "렌탈료_신규": new_df["렌탈료"].sum(min_count=1) if not new_df.empty else 0,
        "렌탈료_연장": ext_df["렌탈료"].sum(min_count=1) if not ext_df.empty else 0,
        "임대대수": len(item_df),
        "임대대수_기존": len(existing_df),
        "임대대수_신규": len(new_df),
        "임대대수_연장": len(ext_df),
        "임대대수_연장대상": len(ext_target_df),
        "자산구분": asset_cnt,
        "당사자산": own_cnt,
        "임차자산": leased_cnt,
        "미분류자산": len(unknown_df),
        "회수율": recovery_rate,
    }

def build_block1(monthly_dfs: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    months = sorted(monthly_dfs.keys(), key=month_end_from_yymm)
    month_label_map = build_month_label_map(months)
    month_labels = [month_label_map[m] for m in months]

    rows = []
    for item in BLOCK1_ITEMS:
        rows.append({"행 레이블": item, **{m: None for m in month_labels}})
        label_order = [
            ("렌탈료", "렌탈료"),
            ("기존", "렌탈료_기존"),
            ("신규", "렌탈료_신규"),
            ("연장", "렌탈료_연장"),
            ("임대대수", "임대대수"),
            ("기존", "임대대수_기존"),
            ("신규", "임대대수_신규"),
            ("연장", "임대대수_연장"),
            ("연장대상", "임대대수_연장대상"),
            ("회수율", "회수율"),
            ("자산구분", "자산구분"),
            ("당사자산", "당사자산"),
            ("임차자산", "임차자산"),
            ("미분류자산", "미분류자산"),
        ]
        monthly_summary = {m: summarize_item_month(monthly_dfs[m], item) for m in months}
        for label, key in label_order:
            row = {"행 레이블": label}
            for m in months:
                row[month_label_map[m]] = monthly_summary[m].get(key, 0)
            rows.append(row)
    return pd.DataFrame(rows)

def build_block2(monthly_dfs: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    months = sorted(monthly_dfs.keys(), key=month_end_from_yymm)
    month_label_map = build_month_label_map(months)
    month_labels = [month_label_map[m] for m in months]

    base_by_month = {m: monthly_dfs[m][monthly_dfs[m]["대상여부"] == 1].copy() for m in months}
    rows = []

    def avg_rental(df_sub: pd.DataFrame):
        cnt = len(df_sub)
        if cnt == 0:
            return "-"
        val = safe_divide(df_sub["렌탈료"].sum(min_count=1), cnt)
        return "-" if pd.isna(val) else val

    for item in BLOCK2_ITEMS:
        rows.append({"구분": item, "세부": None, **{label: None for label in month_labels}})

        ext_rate_row = {"구분": "연장율", "세부": None}
        total_row = {"구분": "렌탈료", "세부": "전체"}
        existing_row = {"구분": "기존", "세부": "기존/대당"}
        new_row = {"구분": "신규", "세부": "신규/대당"}
        ext_row = {"구분": "연장", "세부": "연장/대당"}

        for idx, m in enumerate(months):
            label = month_label_map[m]
            cur_df = base_by_month[m]
            cur_item = cur_df[cur_df["아이템 세분화"] == item].copy()

            total_row[label] = avg_rental(cur_item)
            existing_row[label] = avg_rental(cur_item[cur_item["신규여부"] == "기존"])
            new_row[label] = avg_rental(cur_item[cur_item["신규여부"] == "신규"])
            ext_row[label] = avg_rental(cur_item[cur_item["신규여부"] == "연장"])

            if idx == 0:
                ext_rate_row[label] = "-"
            else:
                prev_m = months[idx - 1]
                prev_df = base_by_month[prev_m]
                prev_ext_target_cnt = int(((prev_df["아이템 세분화"] == item) & (prev_df["연장대상구분"] == "연장대상")).sum())
                cur_extended_cnt = int((cur_item["연장 여부"] == "연장").sum())
                ext_rate_row[label] = "-" if prev_ext_target_cnt == 0 else cur_extended_cnt / prev_ext_target_cnt

        rows.extend([ext_rate_row, total_row, existing_row, new_row, ext_row])
    return pd.DataFrame(rows)

def build_group_dashboard(monthly_dfs: Dict[str, pd.DataFrame], group_col, dashboard_name: str) -> pd.DataFrame:
    months = sorted(monthly_dfs.keys(), key=month_end_from_yymm)
    month_label_map = build_month_label_map(months)

    all_rows = []
    for yymm in months:
        df = monthly_dfs[yymm]
        month_label = month_label_map[yymm]
        base = df[df["대상여부"] == 1].copy()

        agg = (
            base.groupby(group_col, dropna=False)
            .agg(
                임대대수=("코드", "size"),
                월렌탈료=("렌탈료", "sum"),
                취득가=("취득가_num", "sum"),
                신규_count=("신규여부", lambda s: int((s == "신규").sum())),
                연장_count=("신규여부", lambda s: int((s == "연장").sum())),
                당사자산_count=("자산구분_std", lambda s: int((s == "당사자산").sum())),
                임차자산_count=("자산구분_std", lambda s: int((s == "임차자산").sum())),
                미분류자산_count=("자산구분_std", lambda s: int((s == "미분류").sum())),
            )
            .reset_index()
        )

        agg["자산구분_count"] = agg["당사자산_count"] + agg["임차자산_count"]
        agg["대당월렌탈료"] = agg.apply(lambda r: safe_divide(r["월렌탈료"], r["임대대수"]), axis=1)
        agg["회수율"] = agg.apply(lambda r: safe_divide(r["월렌탈료"], r["취득가"]), axis=1)
        agg["월"] = month_label
        all_rows.append(agg)

    long_df = pd.concat(all_rows, ignore_index=True)
    metrics = [
        "임대대수", "월렌탈료", "취득가", "대당월렌탈료", "회수율",
        "신규_count", "연장_count", "자산구분_count",
        "당사자산_count", "임차자산_count", "미분류자산_count"
    ]
    id_cols = group_col if isinstance(group_col, list) else [group_col]

    wide_frames = []
    for metric in metrics:
        p = long_df.pivot_table(index=id_cols, columns="월", values=metric, aggfunc="first")
        p.columns = [f"{metric}_{c}" for c in p.columns]
        wide_frames.append(p)

    out = pd.concat(wide_frames, axis=1).reset_index()
    return out.sort_values(by=id_cols, kind="stable").reset_index(drop=True)

# %% [cell 21 type=markdown]
# ## 9. DEBUG 요약

# %% [cell 22 type=code]
def dict_to_multiline_text(d: Dict) -> str:
    if not d:
        return ""
    return "\n".join([f"{k}: {v}" for k, v in d.items()])

def build_debug_summary(
    raw_by_month: Dict[str, pd.DataFrame],
    stats_by_month: Dict[str, Dict],
    match_rate_by_month: Dict[str, float],
) -> pd.DataFrame:
    months = sorted(raw_by_month.keys(), key=month_end_from_yymm)
    rows = [
        ["RAW 월 목록", ", ".join(months)],
        ["RAW 월 개수", len(months)],
        ["연장 여부 산식", "재계약 여부='연장대상'이면 '연장', 그 외는 '제외'"],
        ["자산구분 산식", "당사자산_count + 임차자산_count"],
    ]

    for idx, m in enumerate(months):
        stats = stats_by_month[m]
        rows.extend([
            [f"{m} raw rows", len(raw_by_month[m])],
            [f"{m} raw cols", len(raw_by_month[m].columns)],
            [f"{m} 대상여부=1 count", stats["대상여부_1_count"]],
            [f"{m} 아이템 NaN 비율", stats["아이템_nan_ratio"]],
            [f"{m} 렌탈료 NaN 비율", stats["렌탈료_nan_ratio"]],
            [f"{m} 취득가 NaN 비율", stats["취득가_nan_ratio"]],
            [f"{m} 코드 중복 개수", stats["코드_중복건수"]],
            [f"{m} 연장키 중복 개수", stats["연장키_중복건수"]],
            [f"{m} 아이템세분화 top20", dict_to_multiline_text(stats["아이템세분화_top20"])],
            [f"{m} 자산구분_std 분포", dict_to_multiline_text(stats["자산구분_std_분포"])],
        ])
        if idx > 0:
            rows.append([f"{m} 연장확인 매칭률", match_rate_by_month.get(m)])

    return pd.DataFrame(rows, columns=["지표", "값"])

# %% [cell 23 type=markdown]
# ## 10. 엑셀 쓰기

# %% [cell 24 type=code]
def write_df_to_sheet(ws, df: pd.DataFrame, start_row: int, start_col: int, write_header: bool = True):
    r = start_row
    c = start_col
    if write_header:
        for j, col_name in enumerate(df.columns, start=c):
            ws.cell(row=r, column=j, value=str(col_name))
        r += 1

    for _, row in df.iterrows():
        for j, col_name in enumerate(df.columns, start=c):
            ws.cell(row=r, column=j, value=num_or_none(row[col_name]))
        r += 1

def save_output_excel(
    out_path,
    block1,
    block2,
    ext_check,
    ext_debug,
    monthly_dfs,
    team_dashboard,
    person_dashboard,
    debug_summary,
    debug_tables_by_month,
):
    wb = Workbook()
    wb.remove(wb.active)

    ws_dash = wb.create_sheet("Dash Board 피벗")
    write_df_to_sheet(ws_dash, block1, start_row=3, start_col=6, write_header=True)

    block2_start_col = max(21, 6 + block1.shape[1] + 2)
    write_df_to_sheet(ws_dash, block2, start_row=3, start_col=block2_start_col, write_header=True)

    ws_ext = wb.create_sheet("연장 확인(파이썬)")
    write_df_to_sheet(ws_ext, ext_check, start_row=1, start_col=1, write_header=True)

    ws_ext_dbg = wb.create_sheet("연장매칭_DEBUG")
    write_df_to_sheet(ws_ext_dbg, ext_debug, start_row=1, start_col=1, write_header=True)

    for yymm in sorted(monthly_dfs.keys(), key=month_end_from_yymm):
        ws_raw = wb.create_sheet(f"{yymm} 정제RAW")
        write_df_to_sheet(ws_raw, monthly_dfs[yymm], start_row=1, start_col=1, write_header=True)

    ws_team = wb.create_sheet("팀별 대시보드")
    write_df_to_sheet(ws_team, team_dashboard, start_row=1, start_col=1, write_header=True)

    ws_person = wb.create_sheet("개인별 대시보드")
    write_df_to_sheet(ws_person, person_dashboard, start_row=1, start_col=1, write_header=True)

    ws_debug = wb.create_sheet("DEBUG_요약")
    write_df_to_sheet(ws_debug, debug_summary, start_row=1, start_col=1, write_header=True)

    all_asset_raw = pd.concat([debug_tables_by_month[m]["asset_raw_dist"] for m in sorted(debug_tables_by_month.keys(), key=month_end_from_yymm)], ignore_index=True)
    all_asset_norm = pd.concat([debug_tables_by_month[m]["asset_norm_dist"] for m in sorted(debug_tables_by_month.keys(), key=month_end_from_yymm)], ignore_index=True)
    all_asset_unknown = pd.concat([debug_tables_by_month[m]["asset_unknown_sample"] for m in sorted(debug_tables_by_month.keys(), key=month_end_from_yymm)], ignore_index=True)

    ws_asset_raw = wb.create_sheet("자산구분_원천분포")
    write_df_to_sheet(ws_asset_raw, all_asset_raw, 1, 1, True)

    ws_asset_norm = wb.create_sheet("자산구분_표준화분포")
    write_df_to_sheet(ws_asset_norm, all_asset_norm, 1, 1, True)

    ws_asset_unknown = wb.create_sheet("자산구분_미분류샘플")
    write_df_to_sheet(ws_asset_unknown, all_asset_unknown, 1, 1, True)

    wb.save(out_path)

# %% [cell 25 type=markdown]
# ## 11. 메인

# %% [cell 26 type=code]
def run_pipeline_all_raw(in_path: str, out_path: str, ref_sheet: str = "기준정보"):
    print_title("Step 1. 기준정보 로드")
    ref_maps = load_reference_maps(in_path, ref_sheet)

    print_title("Step 2. RAW 시트 탐색")
    months = find_all_raw_yymm_sheets(in_path)
    print(f"탐지된 YYMM RAW 시트: {months}")

    print_title("Step 3. RAW 시트 로드 / 검증")
    raw_by_month = {}
    for yymm in months:
        sheet_name = f"{yymm} RAW"
        raw_df = load_raw_sheet(in_path, sheet_name)
        validate_raw_sheet(raw_df, sheet_name)
        raw_by_month[yymm] = raw_df

    print_title("Step 4. 월별 전처리")
    monthly_dfs = {}
    stats_by_month = {}
    debug_tables_by_month = {}
    for yymm in months:
        df, stats, debug_tables = preprocess_month_raw(raw_by_month[yymm], yymm, ref_maps)
        monthly_dfs[yymm] = df
        stats_by_month[yymm] = stats
        debug_tables_by_month[yymm] = debug_tables

    print_title("Step 5. 연장 체인 계산")
    ext_check_list = []
    ext_debug_list = []
    match_rate_by_month = {}

    first_month = months[0]
    monthly_dfs[first_month] = initialize_first_month(monthly_dfs[first_month])

    for idx in range(1, len(months)):
        prev_yymm = months[idx - 1]
        cur_yymm = months[idx]
        cur_df, ext_check, ext_debug, match_rate = apply_extension_chain(
            monthly_dfs[prev_yymm],
            monthly_dfs[cur_yymm],
            cur_yymm,
        )
        monthly_dfs[cur_yymm] = cur_df
        ext_check_list.append(ext_check)
        ext_debug_list.append(ext_debug)
        match_rate_by_month[cur_yymm] = match_rate
        print(f"{prev_yymm} -> {cur_yymm} 연장확인 매칭률: {match_rate:.4%}")

    print_title("Step 6. 연장 / 신규 여부 분포 확인")
    for yymm in months:
        df = monthly_dfs[yymm]
        print(f"\n[{yymm}] 재계약 여부")
        print(df["재계약 여부"].value_counts(dropna=False).to_string())
        print(f"\n[{yymm}] 연장 여부")
        print(df["연장 여부"].value_counts(dropna=False).to_string())
        print(f"\n[{yymm}] 신규여부")
        print(df["신규여부"].value_counts(dropna=False).to_string())

    print_title("Step 7. Dash Board 피벗 생성")
    block1 = build_block1(monthly_dfs)
    block2 = build_block2(monthly_dfs)

    print_title("Step 8. 팀별 / 개인별 대시보드 생성")
    team_dashboard = build_group_dashboard(monthly_dfs, "영업팀", "팀별 대시보드")
    person_dashboard = build_group_dashboard(monthly_dfs, ["영업팀", "영업담당자"], "개인별 대시보드")

    print_title("Step 9. DEBUG_요약 생성")
    debug_summary = build_debug_summary(raw_by_month, stats_by_month, match_rate_by_month)

    print_title("Step 10. 연장 확인 시트용 데이터 생성")
    if ext_check_list:
        ext_check = pd.concat(ext_check_list, ignore_index=True)
        ext_debug = pd.concat(ext_debug_list, ignore_index=True)
    else:
        ext_check = pd.DataFrame(columns=["대상월", "코드", "연장키", "당월_연장대상구분", "전월_연장대상구분_lookup", "당월_연장 여부"])
        ext_debug = pd.DataFrame(columns=["대상월", "코드", "연장키", "자산번호", "BP번호", "주문번호", "청구종료일", "연장대상구분", "재계약 여부", "연장 여부", "신규여부", "전월_연장대상구분", "매칭여부", "연장 여부 원인"])

    print_title("Step 11. 출력 엑셀 저장")
    save_output_excel(
        out_path=out_path,
        block1=block1,
        block2=block2,
        ext_check=ext_check,
        ext_debug=ext_debug,
        monthly_dfs=monthly_dfs,
        team_dashboard=team_dashboard,
        person_dashboard=person_dashboard,
        debug_summary=debug_summary,
        debug_tables_by_month=debug_tables_by_month,
    )
    print(f"저장 완료: {out_path}")

    stats = {
        "months": months,
        "raw_shapes": {m: raw_by_month[m].shape for m in months},
        "대상여부_1": {m: stats_by_month[m]["대상여부_1_count"] for m in months},
        "match_rate_by_month": match_rate_by_month,
        "out_path": out_path,
    }

    print_title("최종 stats")
    print(json.dumps(stats, ensure_ascii=False, indent=2, default=str))
    return stats

# %% [cell 27 type=markdown]
# ## 12. 실행

# %% [cell 28 type=code]
if __name__ == "__main__":
    stats = run_pipeline_all_raw(
        in_path=IN_PATH,
        out_path=OUT_PATH,
        ref_sheet=REF_SHEET,
    )

# %% [cell 29 type=markdown]

